from openpyxl import load_workbook
import os

READ_FILE = "DaytonInformationRajNew10AUG2020.xlsx"
WRITE_FILE = "Packaging_Matrix_UOM_Template_revE - DORMAN USE ONLY.xlsx"
READ_FILE_SHEET = "PKG Data New"
WRITE_FILE_SHEET = "Alt UoM Matrix"


def main():
    output_dir = "output"
    print("Reading {}...".format(READ_FILE))
    data = read_xlsx()
    print("Done")
    print()
    print("Modifying {}...".format(WRITE_FILE))
    create_xlsx(data, output_dir)
    print("Done")
    

def read_xlsx():
    workbook = load_workbook(READ_FILE, read_only=True, data_only=True)
    sheet = workbook[READ_FILE_SHEET]
    data = []
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=8, values_only=True):
        if row[0] is None:
            break
        new_data = dict()
        new_data["partnum"] = row[0]
        new_data["length"] = row[4]
        new_data["width"] = row[5]
        new_data["height"] = row[6]
        new_data["weight"] = row[7]
        data.append(new_data)
    return data

def create_xlsx(data, dir, count=0, num=1):
    workbook = load_workbook(WRITE_FILE, data_only=True)
    sheet = workbook[WRITE_FILE_SHEET]
    for row in sheet.iter_rows(min_row=21, min_col=3, max_col=10):
        if count == len(data):
            break
        if row[0].value == "H":
            row[1].value = data[count]["partnum"]
        elif row[2].value == "EA":
            row[4].value = data[count]["length"]
            row[5].value = data[count]["width"]
            row[6].value = data[count]["height"]
            row[7].value = data[count]["weight"]
            count += 1
    name = "{}/output{}_{}-{}.xlsx".format(dir, num, count-99, count)
    if not os.path.isdir(dir):
        os.mkdir(dir)

    if count < len(data):
        workbook.save(filename=name)
        print("Saved file: {}".format(name))
        create_xlsx(data, dir, count, num+1)
        return
    
    workbook.save(filename=name)
if __name__ == "__main__":
    main()
